#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HGL Reggeli Riport Dashboard Generator
Python verzió - SharePoint Direct Access
"""

import os
import sys
from datetime import datetime, timedelta
import requests
from openpyxl import load_workbook
from io import BytesIO
import json

# ==========================================
# CONFIG
# ==========================================

SHAREPOINT_FILES = {
    'tracking': 'https://hgllog.sharepoint.com/sites/Ecommerce/Megosztott%20dokumentumok/E_COMM%20nyomonk%C3%B6vet%C3%A9s_24.xlsb',
    'ecomm_beosztas': 'https://hgllog.sharepoint.com/sites/Ecommerce/Megosztott%20dokumentumok/E-COM%20%C3%A9s%20TC%20beoszt%C3%A1s/{month}/E-COM+TC%20beoszt%C3%A1s_2026_{month}.xlsx',
    'belepetes': 'https://hgllog.sharepoint.com/sites/Ecommerce/Megosztott%20dokumentumok/Bel%C3%A9ptet%C3%A9s%20H7WH.xlsm',
    'logsupport': 'https://hgllog-my.sharepoint.com/personal/virag_szalai_hgllog_com/Documents/Asztal/műszakbeosztás.xlsx'
}

# ==========================================
# SHAREPOINT AUTH
# ==========================================

def get_access_token():
    """Get Microsoft Graph API access token"""
    tenant_id = os.getenv('SP_TENANT_ID')
    client_id = os.getenv('SP_CLIENT_ID')
    client_secret = os.getenv('SP_CLIENT_SECRET')
    
    if not all([tenant_id, client_id, client_secret]):
        print("⚠️  SharePoint auth nincs beállítva - publikus linkek próbálása...")
        return None
    
    token_url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token'
    token_data = {
        'grant_type': 'client_credentials',
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': 'https://graph.microsoft.com/.default'
    }
    
    response = requests.post(token_url, data=token_data)
    if response.status_code == 200:
        return response.json()['access_token']
    else:
        print(f"❌ Token hiba: {response.status_code}")
        return None

def download_excel(url, token=None):
    """Download Excel file from SharePoint"""
    headers = {}
    if token:
        headers['Authorization'] = f'Bearer {token}'
    
    print(f"📥 Letöltés: {url[:80]}...")
    
    try:
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        print(f"   ✅ OK ({len(response.content) // 1024} KB)")
        return BytesIO(response.content)
    except Exception as e:
        print(f"   ❌ HIBA: {e}")
        return None

# ==========================================
# DATA PROCESSING
# ==========================================

def process_tracking_data(file_obj):
    """Process E_COMM tracking Excel"""
    print("📊 Tracking adatok feldolgozása...")
    
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    
    # TODO: Itt jön a tracking logika (AWB-k, státuszok, stb.)
    # Placeholder adatok most:
    data = {
        'ertesito': 42,
        'felveve': 128,
        'megerkezett': 95,
        'raktar_aktiv': 156,
        'warehouse_pct': 39  # 156t / 400t * 100
    }
    
    wb.close()
    return data

def process_shift_data(file_obj, sheet_name='ECOMM'):
    """Process shift schedule Excel"""
    print(f"📅 {sheet_name} beosztás feldolgozása...")
    
    wb = load_workbook(file_obj, data_only=True)
    
    # TODO: Shift parsing logika
    shifts = []
    
    wb.close()
    return shifts

# ==========================================
# HTML GENERATION
# ==========================================

def generate_html(tracking_data, shifts_ecomm, shifts_belp, shifts_log):
    """Generate dashboard HTML"""
    
    today = datetime.now()
    
    html = f"""<!DOCTYPE html>
<html lang="hu">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HGL Reggeli Riport - {today.strftime('%Y.%m.%d')}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, sans-serif;
            background: #0a0a0a;
            color: #e0e0e0;
            padding: 20px;
        }}
        .header {{
            text-align: center;
            padding: 30px 0;
            border-bottom: 2px solid #333;
        }}
        h1 {{
            color: #4CAF50;
            font-size: 2.5em;
            margin-bottom: 10px;
        }}
        .timestamp {{
            color: #888;
            font-size: 0.9em;
        }}
        .container {{
            max-width: 1400px;
            margin: 40px auto;
        }}
        .card {{
            background: #1a1a1a;
            border-radius: 12px;
            padding: 25px;
            margin-bottom: 30px;
            border: 1px solid #333;
        }}
        .card h2 {{
            color: #4CAF50;
            margin-bottom: 20px;
            font-size: 1.5em;
        }}
        .metrics {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
        }}
        .metric {{
            background: #252525;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}
        .metric-value {{
            font-size: 2.5em;
            color: #4CAF50;
            font-weight: bold;
        }}
        .metric-label {{
            color: #888;
            margin-top: 5px;
            font-size: 0.9em;
        }}
        .progress-bar {{
            height: 30px;
            background: #252525;
            border-radius: 15px;
            overflow: hidden;
            margin: 20px 0;
        }}
        .progress-fill {{
            height: 100%;
            background: linear-gradient(90deg, #4CAF50, #45a049);
            transition: width 0.5s ease;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: bold;
        }}
        .footer {{
            text-align: center;
            margin-top: 50px;
            padding: 20px;
            color: #666;
            border-top: 1px solid #333;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 HGL ECOMMERCE - REGGELI RIPORT</h1>
        <div class="timestamp">Generálva: {today.strftime('%Y. %B %d. %H:%M')}</div>
    </div>
    
    <div class="container">
        <div class="card">
            <h2>📦 E_COMM Nyomonkövetés</h2>
            <div class="metrics">
                <div class="metric">
                    <div class="metric-value">{tracking_data['ertesito']}</div>
                    <div class="metric-label">Értesítő</div>
                </div>
                <div class="metric">
                    <div class="metric-value">{tracking_data['felveve']}</div>
                    <div class="metric-label">Felvéve</div>
                </div>
                <div class="metric">
                    <div class="metric-value">{tracking_data['megerkezett']}</div>
                    <div class="metric-label">Megérkezett</div>
                </div>
                <div class="metric">
                    <div class="metric-value">{tracking_data['raktar_aktiv']}</div>
                    <div class="metric-label">Raktárban aktív</div>
                </div>
            </div>
            
            <h3 style="margin-top: 30px; color: #888;">Raktár kihasználtság (400t kapacitás)</h3>
            <div class="progress-bar">
                <div class="progress-fill" style="width: {tracking_data['warehouse_pct']}%">
                    {tracking_data['warehouse_pct']}% ({tracking_data['raktar_aktiv']}t / 400t)
                </div>
            </div>
        </div>
        
        <div class="card">
            <h2>👥 Mai Beosztások</h2>
            <p style="color: #888;">Beosztás adatok feldolgozás alatt...</p>
        </div>
    </div>
    
    <div class="footer">
        <p>HGL Group Hungary Kft. © {today.year}</p>
        <p>Python Dashboard Generator v1.0 | GitHub Actions</p>
    </div>
</body>
</html>
"""
    
    return html

# ==========================================
# MAIN
# ==========================================

def main():
    print("=" * 50)
    print("HGL REGGELI RIPORT - DASHBOARD GENERÁLÁS")
    print("=" * 50)
    print()
    
    today = datetime.now()
    month = today.strftime('%m')
    
    # 1. Auth
    token = get_access_token()
    
    # 2. Download files
    print("\n📥 FÁJLOK LETÖLTÉSE")
    print("-" * 50)
    
    tracking_file = download_excel(SHAREPOINT_FILES['tracking'], token)
    ecomm_file = download_excel(SHAREPOINT_FILES['ecomm_beosztas'].format(month=month), token)
    belp_file = download_excel(SHAREPOINT_FILES['belepetes'], token)
    log_file = download_excel(SHAREPOINT_FILES['logsupport'], token)
    
    if not tracking_file:
        print("❌ Kritikus hiba: Tracking fájl nem elérhető!")
        sys.exit(1)
    
    # 3. Process data
    print("\n📊 ADATOK FELDOLGOZÁSA")
    print("-" * 50)
    
    tracking_data = process_tracking_data(tracking_file)
    shifts_ecomm = process_shift_data(ecomm_file) if ecomm_file else []
    shifts_belp = process_shift_data(belp_file, 'Beléptetés') if belp_file else []
    shifts_log = process_shift_data(log_file, 'LOG') if log_file else []
    
    # 4. Generate HTML
    print("\n🎨 HTML GENERÁLÁS")
    print("-" * 50)
    
    html = generate_html(tracking_data, shifts_ecomm, shifts_belp, shifts_log)
    
    # 5. Save
    output_path = 'reggeli_riport.html'
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"\n✅ KÉSZ: {output_path}")
    print(f"   Fájlméret: {len(html) // 1024} KB")
    print()

if __name__ == '__main__':
    main()
